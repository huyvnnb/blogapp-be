import os
from typing import Union, Optional

import pytest
import functools
import openpyxl
import unicodedata
from _pytest.outcomes import Failed
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string
import pytest_check as check
import logging

from tests.pages import SafeActionError

logger = logging.getLogger(__name__)
#
# EXCEL_FILE_NAME = "KichBanAutomationTest2.xlsx"
# EXCEL_FILE_PATH = f"{os.getcwd()}/test_case/{EXCEL_FILE_NAME}"
COL_ID = 'B'
COL_DETAIL = 'H'
COL_RESULT = 'I'
COL_EXPECT = 'J'
COL_ACTUAL = 'K'

# Excel style
ALIGNMENT_STYLE = Alignment(wrapText=True, vertical='top')
STATUS_STYLE = Alignment(
    horizontal='center',
    vertical='center',
)
FILL_FAIL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
FILL_PASS = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
FILL_NONE = PatternFill(fill_type=None)

border_all = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

LANG_INDEX_MAP = {
    "vi": 0,
    "en": 1
}


min_row = 4
min_col = column_index_from_string(COL_ID)
max_col = column_index_from_string(COL_EXPECT)


def _normalize(value):
    if isinstance(value, str):
        return unicodedata.normalize('NFC', value.strip())
    elif isinstance(value, list):
        return [_normalize(x) for x in value]
    elif isinstance(value, dict):
        return {_normalize(k): _normalize(v) for k, v in value.items()}
    return _normalize(str(value))


def _check_and_log(actual_value, expected_value, item_name):
    actual_norm = _normalize(actual_value)
    expected_norm = _normalize(expected_value)

    check.equal(
        actual_norm,
        expected_norm,
        f"FAIL: Mục '{item_name}' bị sai. Thực tế: {actual_value}, Mong đợi: {expected_value}"
    )

    return actual_value


def update_excel_result(
        file_path: str,
        test_case_id: str,
        status: str,
        actual_result: str,
        expected_result: str,
        expected_detail: Optional[str] = None
):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        found = False

        for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, values_only=False):
            cell_id = row[0]

            if cell_id.value == test_case_id:
                current_row = cell_id.row

                status_cell = ws[f"{COL_RESULT}{current_row}"]
                status_cell.value = status
                status_cell.alignment = STATUS_STYLE
                status_cell.border = border_all

                if status == 'F':
                    status_cell.fill = FILL_FAIL
                elif status == 'P':
                    status_cell.fill = FILL_PASS
                else:
                    status_cell.fill = FILL_NONE

                actual_cell = ws[f"{COL_ACTUAL}{current_row}"]
                actual_cell.value = actual_result
                actual_cell.alignment = ALIGNMENT_STYLE

                expect_cell = ws[f"{COL_EXPECT}{current_row}"]
                expect_cell.value = expected_result
                expect_cell.alignment = ALIGNMENT_STYLE

                expect_detail_cell = ws[f"{COL_DETAIL}{current_row}"]
                expect_detail_cell.value = expected_detail
                expect_detail_cell.alignment = ALIGNMENT_STYLE

                found = True
                break

        if found:
            wb.save(file_path)
            logger.info(f"Đã cập nhật kết quả cho {test_case_id}.")
        else:
            logger.error(f"Lỗi: Không tìm thấy Test Case ID '{test_case_id}' trong file.")

    except Exception as e:
        logger.error(f"Lỗi khi cập nhật Excel: {e}")


def log_test_result(test_case_ids: Union[str, list[str]]):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            final_test_case_id = None

            if isinstance(test_case_ids, str):
                final_test_case_id = test_case_ids
            elif isinstance(test_case_ids, list):
                current_lang = kwargs.get("lang")
                if not current_lang:
                    raise ValueError(f"Tham số 'lang' bị thiếu khi cung cấp danh sách ID: {test_case_ids}")

                lang_index = LANG_INDEX_MAP.get(current_lang)
                if lang_index is None or lang_index >= len(test_case_ids):
                    raise ValueError(f"Ngôn ngữ/index '{current_lang}' không khớp với list ID: {test_case_ids}.")

                final_test_case_id = test_case_ids[lang_index]

            if not final_test_case_id:
                raise ValueError("Không thể xác định Test Case ID cuối cùng.")

            file_path = kwargs.get("test_report_file")
            if not file_path:
                raise ValueError(
                    f"Hàm test {func.__name__} phải nhận 'test_report_file' làm tham số"
                )

            status = 'F'
            expected_data_parts = []
            actual_data_parts = []
            expected_detail_parts = []
            actual_data = ""
            expected_data = ""
            expected_detail = ""

            try:
                compare_values = func(*args, **kwargs)
                for display_name, expected_val, actual_val in compare_values:
                    expected_detail_parts.append(f"[{display_name}]: {expected_val}")

                    expected_data_parts.append(serialize_data(expected_val))
                    logged_actual_val = _check_and_log(actual_val, expected_val, display_name)
                    actual_data_parts.append(serialize_data(logged_actual_val))

                    # expected_data_parts.append(f"{expected_val}")
                    # logged_actual_val = _check_and_log(actual_val, expected_val, display_name)
                    # actual_data_parts.append(f"{logged_actual_val}")
                status = "P"
            except SafeActionError as e:
                status = "F"
                actual_data = str(e)
                raise e
            except Exception as e:
                status = "F"
                if not actual_data_parts:
                    actual_data = f"LỖI HỆ THỐNG: {str(e)}"
                raise e

            finally:
                if not actual_data:
                    actual_data = '\n'.join(actual_data_parts)
                    expected_data = '\n'.join(expected_data_parts)
                    expected_detail = '\n'.join(expected_detail_parts)

                if check.any_failures():
                    status = "F"

                update_excel_result(
                    file_path=file_path,
                    test_case_id=final_test_case_id,
                    status=status,
                    actual_result=actual_data,
                    expected_result=expected_data,
                    expected_detail=expected_detail
                )

        return wrapper

    return decorator


def compare_set(expect_value: set, actual_value: set):
    matches = expect_value & actual_value
    missing = expect_value - actual_value
    extra = actual_value - expect_value

    result = []
    for item in matches:
        result.append((item, item))

    for item in missing:
        result.append((item, "-"))

    for item in extra:
        result.append(("-", item))

    return result


def serialize_data(value, _format="yaml"):
    if isinstance(value, (dict, list)):
        if _format == "json":
            import json
            return json.dumps(value, ensure_ascii=True, indent=2, sort_keys=True)
        elif _format == "yaml":
            import yaml
            return yaml.dump(value, allow_unicode=True, sort_keys=False, default_flow_style=False)
    return str(value)
