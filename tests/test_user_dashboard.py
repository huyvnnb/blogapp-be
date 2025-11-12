from pathlib import Path

import openpyxl
from playwright.sync_api import Page, expect

from tests.endpoints import Dashboard
from tests.messages import DashboardMsg, DialogMsg
from tests.pages import DashboardPage
from tests.test_user_auth import _login
from tests.testcase_helper import log_test_result
import pytest_check as check


@log_test_result(test_case_ids="TC8.1")
def test_post_status(page: Page, test_factory, base_data, test_report_file):
    user, password = test_factory.create_user(base_data["role_ids"]["user"])

    approve_count = 3
    pending_count = 2
    reject_count = 1
    total_count = approve_count + pending_count + reject_count

    test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="APPROVE",
        count=approve_count
    )

    test_factory.create_posts(
        user,
        title_pref="Pending post",
        summary="Pending summary",
        content="Pending content",
        status="PENDING",
        count=pending_count
    )

    test_factory.create_posts(
        user,
        title_pref="Reject post",
        summary="Reject summary",
        content="Reject content",
        status="REJECT",
        count=reject_count
    )

    _login(page, payload={'username': user.username, 'password': password})

    page.goto(Dashboard.DASHBOARD)
    dashboard_page = DashboardPage(page)

    dashboard_page.post_status_card.is_visible(timeout=5000)
    # actual_approve = dashboard_page.approve_count.inner_text()
    # actual_pending = dashboard_page.pending_count.inner_text()
    # actual_reject = dashboard_page.reject_count.inner_text()
    # actual_total = dashboard_page.total_count.inner_text()
    actual = {
        "status": {
            "approve": int(dashboard_page.approve_count.inner_text()),
            "pending": int(dashboard_page.pending_count.inner_text()),
            "reject": int(dashboard_page.reject_count.inner_text()),
            "total": int(dashboard_page.total_count.inner_text()),
        }
    }

    # --- Dữ liệu mong đợi ---
    expected = {
        "status": {
            "approve": approve_count,
            "pending": pending_count,
            "reject": reject_count,
            "total": total_count,
        }
    }

    return [
        (DashboardMsg.POST_STATISTICS, expected, actual)
    ]


@log_test_result(test_case_ids="TC8.2")
def test_export_data_check_summary_field(page: Page, test_factory, base_data, test_report_file):
    user, password = test_factory.create_user(base_data["role_ids"]["user"])

    _login(page, payload={'username': user.username, 'password': password})
    page.goto(Dashboard.DASHBOARD)

    dashboard_page = DashboardPage(page)

    export_btn = dashboard_page.export_btn
    export_btn.click()

    confirm_export = dashboard_page.confirm_export()
    ids = [
        "user_info",
        "total_posts",
        "avg_length",
        "approval_rate",
    ]
    dashboard_page.select_checkbox(ids)
    with page.expect_download() as download_info:
        confirm_export.click()

    download = download_info.value

    download_dir = Path("./downloads")
    download_dir.mkdir(exist_ok=True)

    path = download.path()
    print(f"Path: {path}")
    file_path = download_dir / download.suggested_filename
    download.save_as(str(file_path))
    print(f"File saved at: {file_path}")

    import os
    compare_values = []
    if not os.path.exists(file_path):
        check.fail("File download thất bại hoặc file trống.")
        return compare_values

    # if os.path.exists(file_path):
    #     compare_values = [(DialogMsg.FILE_FOUND.format(file_name=download.suggested_filename), "", "")]
    # else:
    #     compare_values.append((DialogMsg.FILE_NOT_FOUND.format(file_name=download.suggested_filename), "", ""))
    #     check.fail("File download thất bại hoặc file trống.")
    #     return compare_values

    expected_headers = ["User ID", "Username", "Display name", "Total post", "Average words", "Approval rate"]

    wb = openpyxl.load_workbook(file_path)
    print("Tên các sheet:", wb.sheetnames)
    sheet_names = wb.sheetnames

    summary_sheet = wb[sheet_names[0]]

    summary_data = list(summary_sheet.iter_rows(min_row=1, values_only=True))
    header_row = list(summary_data[0])
    expected = {"summary_header": expected_headers}
    actual = {"summary_header": header_row}

    # compare_values.extend([
    #     (DialogMsg.FIELD_SUMMARY_HEADER, expected_headers, header_row),
    # ])

    if file_path.exists():
        os.remove(file_path)

    return [
        (DialogMsg.FIELD_SUMMARY_HEADER, expected, actual)
    ]


@log_test_result(test_case_ids="TC8.4")
def test_export_data_check_summary_detail(page: Page, test_factory, base_data, test_report_file):
    user, password = test_factory.create_user(base_data["role_ids"]["user"])

    approve_count = 1
    pending_count = 1
    reject_count = 1
    total_count = approve_count + pending_count + reject_count

    all_posts = test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="APPROVE",
        count=approve_count
    )

    all_posts.extend(test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="PENDING",
        count=pending_count
    ))

    all_posts.extend(test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="REJECT",
        count=reject_count
    ))
    _login(page, payload={'username': user.username, 'password': password})
    page.goto(Dashboard.DASHBOARD)

    dashboard_page = DashboardPage(page)

    export_btn = dashboard_page.export_btn
    export_btn.click()

    confirm_export = dashboard_page.confirm_export()
    ids = [
        "user_info",
        "total_posts",
        "avg_length",
        "approval_rate",
    ]
    dashboard_page.select_checkbox(ids)
    with page.expect_download() as download_info:
        confirm_export.click()

    download = download_info.value

    download_dir = Path("./downloads")
    download_dir.mkdir(exist_ok=True)

    path = download.path()
    print(f"Path: {path}")
    file_path = download_dir / download.suggested_filename
    download.save_as(str(file_path))
    print(f"File saved at: {file_path}")

    import os
    compare_values = []
    if not os.path.exists(file_path):
        check.fail("File download thất bại hoặc file trống.")
        return compare_values

    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames

    summary_sheet = wb[sheet_names[0]]

    summary_data = list(summary_sheet.iter_rows(min_row=1, values_only=True))
    summary_row = list(summary_data[1])

    uid, username, display_name, total_post, avg_words, approval_rate = summary_row
    total_words = sum(len(post.content.split()) for post in all_posts)
    expected_avg = total_words // len(all_posts) if all_posts else 0
    expected_approval = approve_count * 100 / total_count if total_count else 0

    expected_summary = {
        "user_id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "total_posts": total_count,
        "avg_words": expected_avg,
        "approval_rate": f"{expected_approval:.2f}%"
    }

    actual_summary = {
        "user_id": uid,
        "username": username,
        "display_name": display_name,
        "total_posts": total_post,
        "avg_words": int(avg_words),
        "approval_rate": approval_rate
    }

    expected = {"summary_detail": expected_summary}
    actual = {"summary_detail": actual_summary}

    # compare_values.append((DialogMsg.FIELD_SUMMARY_DATA, expected_summary, actual_summary))

    if file_path.exists():
        os.remove(file_path)

    return [(DialogMsg.FIELD_SUMMARY_DATA, expected, actual)]


@log_test_result(test_case_ids="TC8.3")
def test_export_data_check_posts_field(page: Page, test_factory, base_data, test_report_file):
    user, password = test_factory.create_user(base_data["role_ids"]["user"])

    approve_count = 1

    test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="APPROVE",
        count=approve_count
    )

    _login(page, payload={'username': user.username, 'password': password})
    page.goto(Dashboard.DASHBOARD)

    dashboard_page = DashboardPage(page)

    export_btn = dashboard_page.export_btn
    export_btn.click()

    confirm_export = dashboard_page.confirm_export()
    ids = [
        "export_posts",
        "post_id",
        "post_title",
        "post_content",
        "post_summary",
        "post_user_id",
        "post_published",
        "post_is_public",
        "post_status",
        "post_created_at",
        "post_updated_at",
    ]
    dashboard_page.select_checkbox(ids)
    with page.expect_download() as download_info:
        confirm_export.click()

    download = download_info.value

    download_dir = Path("./downloads")
    download_dir.mkdir(exist_ok=True)

    file_path = download_dir / download.suggested_filename
    download.save_as(str(file_path))

    import os
    compare_values = []
    if not os.path.exists(file_path):
        check.fail("File download thất bại hoặc file trống.")
        return compare_values

    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames

    posts_sheet = wb[sheet_names[1]]

    posts_data = list(posts_sheet.iter_rows(min_row=1, values_only=True))
    expected_post_headers = ["id", "title", "content", "summary", "user_id", "published", "is_public", "status", "created_at", "updated_at"]
    actual_post_headers = list(posts_data[0])

    expected = {"posts_header": expected_post_headers}
    actual = {"posts_header": actual_post_headers}

    # compare_values.append(
    #     (DialogMsg.POST_DETAIL_HEADER, expected_post_headers, actual_post_headers)
    # )

    if file_path.exists():
        os.remove(file_path)

    return [
        (DialogMsg.POST_DETAIL_HEADER, expected, actual)
    ]


@log_test_result(test_case_ids="TC8.5")
def test_export_data_check_posts_detail(page: Page, test_factory, base_data, test_report_file):
    user, password = test_factory.create_user(base_data["role_ids"]["user"])

    approve_count = 1
    pending_count = 1
    reject_count = 1

    all_posts = test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="APPROVE",
        count=approve_count
    )

    all_posts.extend(test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="PENDING",
        count=pending_count
    ))

    all_posts.extend(test_factory.create_posts(
        user,
        title_pref="Approve post",
        summary="Approve summary",
        content="Approve content",
        status="REJECT",
        count=reject_count
    ))
    _login(page, payload={'username': user.username, 'password': password})
    page.goto(Dashboard.DASHBOARD)

    dashboard_page = DashboardPage(page)

    export_btn = dashboard_page.export_btn
    export_btn.click()

    confirm_export = dashboard_page.confirm_export()
    ids = [
        "export_posts",
        "post_id",
        "post_title",
        "post_content",
        "post_summary",
        "post_user_id",
        "post_published",
        "post_is_public",
        "post_status",
        "post_created_at",
        "post_updated_at",
    ]
    dashboard_page.select_checkbox(ids)
    with page.expect_download() as download_info:
        confirm_export.click()

    download = download_info.value

    download_dir = Path("./downloads")
    download_dir.mkdir(exist_ok=True)

    file_path = download_dir / download.suggested_filename
    download.save_as(str(file_path))

    import os
    compare_values = []
    if not os.path.exists(file_path):
        check.fail("File download thất bại hoặc file trống.")
        return compare_values

    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames

    posts_sheet = wb[sheet_names[1]]

    posts_data = list(posts_sheet.iter_rows(min_row=1, values_only=True))

    expected = {"posts_detail": []}
    actual = {"posts_detail": []}

    for post in all_posts:
        post_dict_expected = {
            "id": post.id,
            "title": post.title,
            "content": post.content,
            "summary": post.summary,
            "user_id": post.user_id,
            "published": post.published,
            "is_public": post.is_public,
            "status": str(post.status),
            "created_at": post.created_at.replace(microsecond=0),
            "updated_at": post.updated_at.replace(microsecond=0),
        }
        expected["posts_detail"].append(post_dict_expected)

    for row in posts_data[1:]:
        post_dict_actual = {
            "id": row[0],
            "title": row[1],
            "content": row[2],
            "summary": row[3],
            "user_id": row[4],
            "published": row[5],
            "is_public": row[6],
            "status": row[7],
            "created_at": row[8].replace(microsecond=0),
            "updated_at": row[9].replace(microsecond=0),
        }
        actual["posts_detail"].append(post_dict_actual)

    compare_values.append((DialogMsg.POST_DETAIL_DATA, expected, actual))

    if file_path.exists():
        os.remove(file_path)

    return compare_values
